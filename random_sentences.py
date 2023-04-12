import nltk
from nltk.corpus import treebank
from openpyxl import Workbook

import random

import os
# import torch
# from torchtext.datasets import WikiText2
# from torchtext.data.utils import get_tokenizer

# tokenizer = get_tokenizer('basic_english')
# dataset = WikiText2(root='data')

# random_short_sentences = []
# for doc in dataset:
#     for sentence in tokenizer(doc.text):
#         words = sentence.split()
#         if len(random_short_sentences) <= 18:
#             random_short_sentences.append(sentence)
#         if len(random_short_sentences) == 2000:
#             break
#     if len(random_short_sentences) == 2000:
#         break

# download The Penn Treebank dataset if not already downloaded
nltk.download('treebank')

# get a list of all sentence tokens in The Penn Treebank
sentences = treebank.sents()

# Filter sentences to only include ones with length less than 18 words
short_sentences = [s for s in sentences if len(s) < 27]

# randomly select 2000 sentences from the list of all sentences
random_sentences = random.sample(sorted(sentences), 2000)

# Choose 2000 random short sentences
# random_short_sentences = random.sample(sorted(short_sentences), 2000)
if len(short_sentences) >= 2000:
    random_short_sentences = random.sample(short_sentences, 2000)
else:
    random_short_sentences = short_sentences
for sentence in random_short_sentences:
    for word in sentence:
        if "*" in word:
            sentence.remove(word)
# create a new Excel workbook and worksheet
wb = Workbook()
ws = wb.active

# # add each sentence to a new row in the worksheet
# for i, sentence in enumerate(random_sentences):
#     ws.cell(row=i+1, column=1, value=' '.join(sentence))
# add each sentence to a new row in the worksheet
for i, sentence in enumerate(random_short_sentences):
    ws.cell(row=i+1, column=1, value=' '.join(sentence))

# # save the workbook to a new Excel file
# wb.save("random_penn_treebank_sentences.xlsx")
# # print(os.getcwd())
# try:
#     wb.save("random_penn_treebank_sentences.xlsx")
#     print("File saved successfully")
# except Exception as e:
#     print("Error saving file:", e)

# save the workbook to a new Excel file
# wb.save("random_penn_treebank_short_sentences.xlsx")
# print(os.getcwd())
try:
    wb.save("random_penn_treebank_short_sentences.xlsx")
    print("File saved successfully")
except Exception as e:
    print("Error saving file:", e)

# # save the workbook to a new Excel file
# wb.save("random_wikitext2_short_sentences.xlsx")
print(os.getcwd())
# try:
#     wb.save("random_wikitext2_short_sentences.xlsx")
#     print("File saved successfully")
# except Exception as e:
#     print("Error saving file:", e)



# for i, sentence in enumerate(random_short_sentences):
#     ws.cell(row=i+1, column=1, value=' '.join(sentence))

# try:
#     wb.save("random_penn_treebank_short_sentences.xlsx")
#     print("File saved successfully")
# except Exception as e:
#     print("Error saving file:", e)



# # check if the file exists in the current working directory
# if "random_penn_treebank_sentences.xlsx" in os.listdir():
#     print("File found in directory")
# else:
#     print("File not found in directory")   

# import os
# print(os.getcwd())  # print the current working directory

# # check if the file exists in the current working directory
# if "random_wikitext2_short_sentences.xlsx" in os.listdir():
#     print("File found in directory")
# else:
#     print("File not found in directory")    
# # print(os.getcwd())